"""
XLSX report builder for survey analysis.
Creates structured Excel workbooks with cross-tabs, statistics, and formatting.
"""

import logging
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Styles
HEADER_FILL = PatternFill(start_color='2E5B88', end_color='2E5B88', fill_type='solid')
HEADER_FONT = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
ALT_FILL = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid')
NORMAL_FONT = Font(name='Calibri', size=9)
BOLD_FONT = Font(name='Calibri', size=9, bold=True)
THIN_BORDER = Border(
    left=Side(style='thin', color='D0D5DD'),
    right=Side(style='thin', color='D0D5DD'),
    top=Side(style='thin', color='D0D5DD'),
    bottom=Side(style='thin', color='D0D5DD'),
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)


class XlsxReportBuilder:
    """Builds structured Excel workbooks for survey reports."""
    
    def __init__(self):
        self.wb = Workbook()
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
    
    def add_dataframe_sheet(self, df: pd.DataFrame, sheet_name: str,
                            title: str = "", auto_width: bool = True):
        """Add a DataFrame as a formatted sheet."""
        # Sanitize sheet name (max 31 chars, no special chars)
        safe_name = sheet_name[:31].replace('/', '-').replace('\\', '-')
        safe_name = safe_name.replace('[', '(').replace(']', ')').replace('*', '')
        safe_name = safe_name.replace('?', '').replace(':', '-')
        
        ws = self.wb.create_sheet(title=safe_name)
        
        start_row = 1
        if title:
            ws.cell(row=1, column=1, value=title)
            ws.cell(row=1, column=1).font = Font(name='Calibri', size=12, bold=True, color='2E5B88')
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
            start_row = 3
        
        # Headers
        for j, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=j, value=str(col_name))
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER
        
        # Data
        for i, row_data in enumerate(df.itertuples(index=False), start_row + 1):
            for j, val in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=j)
                if pd.isna(val):
                    cell.value = ''
                elif isinstance(val, float):
                    cell.value = round(val, 2)
                    cell.number_format = '0.00'
                else:
                    cell.value = val
                
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
                
                if j == 1:
                    cell.alignment = LEFT_WRAP
                else:
                    cell.alignment = CENTER
                
                # Alternating rows
                if (i - start_row) % 2 == 0:
                    cell.fill = ALT_FILL
        
        # Auto-width
        if auto_width:
            for j in range(1, len(df.columns) + 1):
                col_letter = get_column_letter(j)
                max_len = max(
                    len(str(df.columns[j-1])),
                    *[len(str(df.iloc[i, j-1])) for i in range(min(len(df), 50))]
                )
                ws.column_dimensions[col_letter].width = min(max(max_len * 1.2 + 2, 8), 50)
        
        # Freeze header
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
        
        return ws
    
    def add_cross_tab_sheet(self, cross_tab: pd.DataFrame, sheet_name: str,
                            title: str = ""):
        """Add a cross-tabulation as a formatted sheet."""
        # Reset index to make it a regular DataFrame
        ct = cross_tab.reset_index()
        ct.columns = [str(c) for c in ct.columns]
        return self.add_dataframe_sheet(ct, sheet_name, title)
    
    def save(self, filepath: str):
        """Save workbook."""
        if not self.wb.sheetnames:
            self.wb.create_sheet('Empty')
        self.wb.save(filepath)
        logger.info(f"XLSX report saved to {filepath}")
